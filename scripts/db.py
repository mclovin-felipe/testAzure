import openpyxl
import os
import pandas as pd
import json
os.chdir(os.path.dirname(os.path.realpath(__file__)))
file = 'TI.xlsx'
data = pd.ExcelFile(file)

df = data.parse('Detalle RRHH TI 2023')
# df.info()
# df.head()
ps = openpyxl.load_workbook('TI.xlsx')
sheet = ps['Detalle RRHH TI 2023']
personas = []
for row in range(2, sheet.max_row + 1):
    # print(row)
    persona = {
        "id": row,
        "nombre": sheet.cell(row, 8).value,
        "equipo": sheet.cell(row, 6).value,

    }
    # dict[sheet.cell(row, 1).value] = persona
    print(persona)
    personas.append(persona)
personas = json.dumps(personas, ensure_ascii=False, indent=4)
file = open('person.json', 'w')
file.write(personas)
print(personas)
